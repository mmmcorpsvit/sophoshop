import os
# from os import rename
# import tarfile
# import zipfile
# import zlib

import shutil
import tempfile
import urllib3

import logging

from django.core.files import File
from django.core.management.base import BaseCommand, CommandError
from django.core import management
# from oscar.core.loading import get_class

# from decimal import Decimal as D

from django.db.transaction import atomic
# from django.db import transaction
# from django.utils.translation import ugettext_lazy as _
# from django.db.transaction import atomic

from oscar.apps.catalogue.categories import create_from_breadcrumbs
from oscar.core.loading import get_class, get_classes, get_model

from openpyxl import load_workbook

# from shutil import move
from PIL import Image

import subprocess

from settings import SITE_ROOT

ImportingError = get_class('partner.exceptions', 'ImportingError')
Partner, StockRecord = get_classes('partner.models', ['Partner', 'StockRecord'])

ProductClass, ProductAttribute, Product, Category, ProductCategory, ProductAttributeValue = get_classes(
    'catalogue.models', ('ProductClass', 'ProductAttribute', 'Product',
                         'Category', 'ProductCategory', 'ProductAttributeValue'))


AttributeOption, AttributeOptionGroup = get_classes(
    'catalogue.models', ('AttributeOption', 'AttributeOptionGroup'))

ProductImage = get_model('catalogue', 'productimage')


logger = logging.getLogger('oscar.catalogue.import')
urllib3.disable_warnings()

# use: sophoshop_import_from_xls_prom export-products.xlsx --flush --rebuild_index --add_images
# must be in _private/ImageMagic/convert.exe from ImagePagic Portable


def run_win_cmd(cmd):
    result = []
    process = subprocess.Popen(cmd,
                               shell=True,
                               stdout=subprocess.PIPE,
                               stderr=subprocess.PIPE)
    for line in process.stdout:
        result.append(line)
    errcode = process.returncode
    for line in result:
        logger.info(line)
    if errcode is not None:
        logger.error(cmd)
        raise Exception('cmd %s failed, see above for details', cmd)


def download(c, url, filename):
    """
    Download file
    :param c: c = urllib3.PoolManager()
    :param url: URL
    :param filename: filename where file saved
    # :return: data of file
    """

    # logger.error(url)
    # logger.error(filename)

    with c.request('GET', url, preload_content=False) as resp, open(filename, 'wb') as out_file:
        shutil.copyfileobj(resp, out_file)

    resp.release_conn()  # not 100% sure this is required though
    # return data


class Impxls(object):
    _flush = False
    _add_images = False
    _rebuild_index = False
    _csv1 = None
    _csv2 = None

    def __init__(self, logger2, flush=False, add_images=False, rebuild_index=False):
        self.logger = logger2
        self._flush = flush
        self._add_images = add_images
        self._rebuild_index = rebuild_index
        self._csv1 = open('attr1.csv', 'w')
        self._csv2 = open('attr2.csv', 'w')

    def _create_item(self, row, product_class_name, brand, country_manufactur,
                     category_str, upc, title, description, images_urls,
                     price, attr_list, stats):
        # Ignore any entries that are NULL
        if description == 'NULL':
            description = ''

        # Create item class and item
        # product_class, __ = ProductClass.objects.get_or_create(name=product_class_name, track_stock=False)
        klass, __ = ProductClass.objects.get_or_create(name=product_class_name, track_stock=False)

        # from Solr search, schema.xml
        # <dynamicField name="*_i"  type="int"    indexed="true"  stored="true"/>
        # <dynamicField name="*_s"  type="string"  indexed="true"  stored="true"/>
        # <dynamicField name="*_l"  type="long"   indexed="true"  stored="true"/>
        # <dynamicField name="*_t"  type="text_en"    indexed="true"  stored="true"/>
        # <dynamicField name="*_b"  type="boolean" indexed="true"  stored="true"/>
        # <dynamicField name="*_f"  type="float"  indexed="true"  stored="true"/>
        # <dynamicField name="*_d"  type="double" indexed="true"  stored="true"/>
        # <dynamicField name="*_dt" type="date" indexed="true" stored="true"/>
        # <dynamicField name="*_p" type="location" indexed="true" stored="true"/>
        # <dynamicField name="*_coordinate"  type="tdouble" indexed="true"  stored="false"/>

        ProductAttribute.objects.get_or_create(  # create and/or get created class
            product_class=klass,
            name='Виробник',    # text in admin
            required=False,
            code='brand_s',    # name in DB (use bootom connemt sufix)
            type='text',        # type
            )

        ProductAttribute.objects.get_or_create(
            product_class=klass,
            name='Країна виробник',
            required=False,
            code='country_manufactur_s',
            type='text',
            )

        # dymanic create attributes

        item = Product(product_class=klass,  # inherit from class!
                       title=title,
                       upc=upc,
                       description=description,
                       )
        # if not (price is None):
        #     item.price = price

        item.price = price

        # Set attributes
        item.attr.brand_s = brand
        item.attr.country_manufactur_s = country_manufactur

        item.save()

        for el in attr_list:
            pass


        # Associate with a category
        cat = create_from_breadcrumbs(category_str)
        ProductCategory.objects.update_or_create(product=item, category=cat)

        # Set the price
        self._create_stockrecord(item, 'Склад по змовчуванню', upc, price)  # use one stock how main for sales

        # region 'image'
        c = urllib3.PoolManager()
        if self._add_images:
            images = str(images_urls).split(',')
            for image in images:
                image_url = image.strip()
                if len(image_url) < 5:
                    continue

                file_name = image.replace('https://images.ua.prom.st/', '').strip()
                fn = tempfile.gettempdir() + '\\' + file_name
                fn = fn.strip()

                download(c, image_url, fn)

                # fix #15, some files has png on jpeg file error content
                with Image.open(fn) as img:
                    image_format = img.format

                # logger.info(img.format)  # 'JPEG'
                if image_format == 'PNG':
                    nfn = os.path.splitext(fn)[0]+'.jpg'
                    s = '%s/%s/convert.exe "%s" -background white -flatten "%s"' % \
                        (SITE_ROOT, '_private/ImageMagic', fn, nfn)
                    logger.info(s)
                    res = run_win_cmd(s)
                    logging.info(res)
                    fn = nfn
                else:
                    if not image_format == 'JPEG':
                        logger.error('image_format=%s' % image_format)
                        exit()

                new_file = File(open(fn, 'rb'))
                im = ProductImage(product=item)
                im.original.save(file_name, new_file, save=False)
                im.save()
                logger.debug('Image added to "%s"' % item)
        # endregion

        return item

    @staticmethod
    def _create_stockrecord(item, partner_name, partner_sku, price):
        def d(x):
            return int(x)

        # Create partner and stock record
        partner, _ = Partner.objects.get_or_create(name=partner_name)
        try:
            stock = StockRecord.objects.get(partner_sku=partner_sku)
        except StockRecord.DoesNotExist:
            stock = StockRecord()

        stock.product = item
        stock.partner = partner
        stock.partner_sku = partner_sku
        stock.price_excl_tax = d(price)
        stock.price_retail = d(price)
        # stock.num_in_stock = num_in_stock
        stock.save()

    @atomic
    def _flush_product_data(self):
        """Flush out product and stock models"""
        logger.info('Flush start')
        ProductCategory.objects.all().delete()
        Category.objects.all().delete()
        Product.objects.all().delete()
        ProductClass.objects.all().delete()
        ProductAttribute.objects.all().delete()
        Partner.objects.all().delete()
        StockRecord.objects.all().delete()
        AttributeOptionGroup.objects.all().delete()
        AttributeOption.objects.all().delete()
        if not self._add_images:
            logger.info('Flush images')
            ProductImage.objects.all().delete()
        logger.info('Flush end')

    def handle(self, fn):
        if self._flush:
            self._flush_product_data()

        stats = {'new_items': 0,
                 'updated_items': 0}

        wb2 = load_workbook(fn)
        logger.info('Spread sheats names: %s' % wb2.get_sheet_names())
        wb = wb2.worksheets[0]

        skip_first_row = True
        index = 0
        cats = dict()

        for row in wb.rows:
            index += 1

            # if index > 50:
            #     break

            # if index < 71:
            #    continue

            if skip_first_row:
                logger.info('[0 = skip]')
                skip_first_row = False
                continue

            # logger.info(skip_first_row)
            v = row[5].value
            price = int(v) if not (v is None) else 0

            v = row[3].value
            description = str(v) if not (v is None) else ''

            cat = str(row[15].value) \
                .replace('Матраци Sleep&Fly', 'Матраци') \
                .replace('Матраци Evolution', 'Матраци') \
                .replace('Матраци Sleep&fly Organic', 'Матраци') \
                .replace('Матраци Take&go Bamboо', 'Матраци') \
                .replace('Матраци Sleep&fly uno', 'Матраци') \
                .replace('Матраци на дивани', 'Футони і топери') \
                .replace('Наматрацникии', 'Наматрацники і підматрацники') \
                .replace("Дерев'яні ліжка", 'Ліжка') \
                .replace("Дитячі ліжка", 'Ліжка') \
                .replace('Столи', 'Столи гостьові') \
                .replace('Столи гостьові-трансформери', 'Столи журнальні')\
                .replace('Стільці', 'Стільці та табурети') \
                .replace('Дитячі дивани', 'Дивани') \
                .replace('Кутові дивани', 'Дивани') \
                .replace('Прямі дивани', 'Дивани')
            cats[cat] = None

            brand = str(row[24].value) \
                .replace('Скиф', 'Скіф') \
                .replace('Тиса мебель', 'Тиса меблі') \
                .replace('Елисеевская мебель', 'Єлисеївські меблі') \
                .replace('Микс мебель', 'Мікс меблі') \
                .replace('Мелитополь мебель', 'Мелітополь меблі')

            country_manufactur = str(row[26].value) \
                .replace('Украина', 'Україна')

            some_list = row[30:]

            # get list of tuples(name, value)
            # use step by 1
            attrs_list = [(str(row[1].value),
                           str(val.value),
                           str(some_list[idx+1].value),
                           str(some_list[idx+2].value))
                          for idx, val in enumerate(some_list)
                          if idx % 3 == 0 and val.value is not None and some_list[idx+2].value is not None]

            for e in attrs_list:
                s = e[0]+'\t'+e[1]+'\t'+e[2] + '\t'+e[3] + '\n'
                self._csv1.writelines(s)
                # self._csv += e[0]+'\t'+e[1]+'\t'+e[2]+'\n'

            # self._create_item(
            #     # product_class=str(row[16].value).replace('https://prom.ua/', ''),
            #     row,
            #     product_class_name=cat,
            #     brand=brand,
            #     country_manufactur=country_manufactur,
            #     category_str=cat,
            #     upc=str(row[20].value),
            #     title=str(row[1].value).strip(),
            #     description=description,
            #     price=price,
            #     stats=stats,
            #     images_urls=str(row[11].value),
            #     attr_list=attrs_list,
            # )

            # logger.info('[%i/%i] %s ' % (row.index, wb.rows.count, row[1].value))
            logger.info('[%i/%i] [%s] %s' % (index - 1, wb.max_row, cat, row[1].value,))
            # logger.info(row[1].value)

        # call rebuild search index
        if self._rebuild_index:
            management.call_command('rebuild_index', '--noinput', verbosity=0, interactive=False)

        msg = "New items: %d, updated items: %d" % (stats['new_items'],
                                                    stats['updated_items'])
        # self.logger.info(msg)
        self.logger.info(cats)


class Command(BaseCommand):
    help = 'Import Products and Categories from Prom.ua exported XLS file'

    def add_arguments(self, parser):
        parser.add_argument(
            'filename', nargs='+',
            help='/path/to/file1.xls /path/to/file2.xls ...')
        parser.add_argument(
            '--flush',
            action='store_true',
            dest='flush',
            default=False,
            help='Flush tables before importing')

        parser.add_argument(
            '--add_images',
            action='store_true',
            dest='add_images',
            default=False,
            help='Process images importing')

        parser.add_argument(
            '--rebuild_index',
            action='store_true',
            dest='rebuild_index',
            default=False,
            help='Rebuild search engine index process')

    def handle(self, *args, **options):
        logger.info("Starting catalogue import")

        for file_path in options['filename']:
            logger.info(" - Importing records from '%s'" % file_path)
            try:
                xls = Impxls(logger, flush=options.get('flush'), add_images=options.get('add_images'),
                             rebuild_index=options.get('rebuild_index'))
                xls.handle(file_path)

            except ImportingError as e:
                raise CommandError(str(e))
