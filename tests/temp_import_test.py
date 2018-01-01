# !/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from environs import Env
import xmlrpc.client as xmlrpclib
import base64
import copy
import os
import urllib3
import shutil
import subprocess
from PIL import Image

# import pickle


env = Env()
out = print
urllib3.disable_warnings()


class Utils:
    @staticmethod
    def cls():
        os.system('cls' if os.name == 'nt' else 'clear')

    @staticmethod
    def run_win_cmd(cmd_line):
        result = []
        process = subprocess.Popen(cmd_line,
                                   shell=True,
                                   stdout=subprocess.PIPE,
                                   stderr=subprocess.PIPE)
        out(cmd_line)
        for line in process.stdout:
            result.append(line)
        errcode = process.returncode
        for line in result:
            out(line)
        if errcode is not None:
            out(cmd_line)
            raise Exception('cmd %s failed, see above for details', cmd_line)

    @staticmethod
    def get_base(file_name):
        result = ''
        with open(file_name, 'rb') as resp:
            file_data = resp.read()
            result = base64.encodebytes(file_data).decode("utf-8")
        return result

    @staticmethod
    def get_image_base64_from_url(urlobject, url):
        with urlobject.request('GET', url, preload_content=False) as resp:
            result = base64.b64encode(resp.data)
        resp.release_conn()  # not 100% sure this is required though
        return result

    @staticmethod
    def check_image_correct(fn):
        # test open
        result = False
        try:
            img = Image.open(fn)
            format_str = img.format  # NOQA
            result = True
        except IOError:
            s = ''
            # out('***Error: image open fail: %s, %s ***' % (url, fs))
        return result


class ImportToOdd:
    _srv, _db, _username, _password = '', '', '', ''
    _uid = None
    _models_objects = None
    _common_objects = None
    _cats = dict()

    def __init__(self, _srv, _db, _username, _password):
        self._srv, self._db, self._username, self._password = _srv, _db, _username, _password
        # self._cats = cats
        self.connect()

    def connect(self):
        self._cats.clear()
        rpc = '%s/xmlrpc/2/' % self._srv

        try:
            self._common_objects = xmlrpclib.ServerProxy('%scommon' % rpc)
            cv = self._common_objects.version()
            out('\n==server info==')
            out(cv)

            self._uid = self._common_objects.authenticate(self._db, self._username, self._password, {})
        except ConnectionRefusedError:
            pass

        if not self._uid:
            out('Authorization FAILED!')
            exit(-1)

        out('Authorization succes, uid: %s' % self._uid)
        self._models_objects = xmlrpclib.ServerProxy('%sobject' % rpc)

    def create_category(self, cat_name):
        """
        ReCreate Categories list in root (use internal _self.cats for speed)
        :param cat_name: list of names categories
        :return: id
        """
        # already present in list?
        if cat_name in self._cats:
            return self._cats[cat_name]

        # get cats list
        categories = self._models_objects.execute_kw(
            self._db, self._uid, self._password,
            'product.public.category',  # model (just see param model in admin side URL)
            'search_read',  # operation
            [[
                # conditions
                ['parent_id', 'in', [0, ]],  # one item
                ['name', 'in', [cat_name]],  # item in set
            ],
                ['id']  # fields list
            ]
        )

        # categories = [1, 2]
        # category dont exists, create him
        if len(categories) > 0:
            tid = categories[0]['id']
            self._cats[cat_name] = tid
            return tid

        # create
        result = self._models_objects.execute_kw(
            self._db, self._uid, self._password,
            'product.public.category',

            'create',  # delete
            [{'name': cat_name}]  # from list of id's
        )
        out('Create category: %s, id: %i' % (cat_name, result))
        return result

    def unlink_item(self):
        l2 = [10, 11, 20, 26, 21]

        # check if the deleted record is still in the database
        lst3 = self._models_objects.execute_kw(
            self._db, self._uid, self._password,
            'product.template',
            'search',
            [
                [
                    ['id', 'not in', l2]
                ]
            ]
        )

        for e2 in lst3:
            lst2 = self._models_objects.execute_kw(
                self._db, self._uid, self._password,
                'product.template',
                'unlink',
                [e2])

        out('\n==unlink items==')
        out('count: %i' % len(lst3))
        pass

    def unlink_attributes(self):
        l2 = [1, 2, 3]

        attribute_id = self._models_objects.execute_kw(
            self._db, self._uid, self._password,
            'product.attribute',  # model (just see param model in admin side URL)
            'search_read',  # operation
            [[
                # conditions
                # ['parent_id', 'in', [False, ]],  # one item
                ['id', 'not in', l2],  # item in set
            ],

                ['id']  # fields list
            ]
        )

        for e2 in attribute_id:
            lst2 = self._models_objects.execute_kw(
                self._db, self._uid, self._password,
                'product.attribute',
                'unlink',
                [e2])

        pass

    def create_attribute(self, item, attrs_lines, sname, svalue, price=None):
        if svalue == 'None':
            if sname in ['Бренд']:
                svalue = 'Україна'
            if sname in ['Країна виробник']:
                svalue = 'Україна'

            # _ = 1

        # region 'attribute_id'
        attribute_id = self._models_objects.execute_kw(
            self._db, self._uid, self._password,
            'product.attribute',  # model (just see param model in admin side URL)
            'search_read',  # operation
            [[
                ['name', 'in', [sname]],  # item in set
            ],
                ['id']  # fields list
            ]
        )

        # create attribute if not exists
        if not attribute_id:
            attribute_id = self._models_objects.execute_kw(
                self._db, self._uid, self._password,
                'product.attribute',
                'create',
                [{
                    'name': sname,
                }]
            )
            # attribute_id = attribute_id[0]['id']
        else:
            attribute_id = attribute_id[0]['id']
        # endregion

        # region 'attribute value'
        value_attribute_id = self._models_objects.execute_kw(
            self._db, self._uid, self._password,
            'product.attribute.value',  # model (just see param model in admin side URL)
            'search_read',  # operation
            [[
                # conditions
                # ['parent_id', 'in', [False, ]],  # one item
                ['attribute_id', 'in', [attribute_id]],
                ['name', 'in', [svalue]],  # item in set
            ],
                ['id']  # fields list
            ]
        )

        # create attribute value if not exists
        if not value_attribute_id:
            value_attribute_id = self._models_objects.execute_kw(
                self._db, self._uid, self._password,
                'product.attribute.value',
                'create',
                [{
                    'attribute_id': attribute_id,
                    'name': svalue,
                    # 'html_color': False,
                    'type': 'select',
                }]
            )
        else:
            value_attribute_id = value_attribute_id[0]['id']

        attrs_lines += [
            (0, 0,  # what it is?
             {'attribute_id': attribute_id,  # attribute id
              'value_ids': [(4, value_attribute_id), ]},  # [(unknown ???, value_ids)]
             ),
        ]
        # endregion

        return attribute_id, value_attribute_id

    def create_item(self, item):
        """
        Create in from lst
        :param item: list of dict(name, cat_name, price, desc)
        :return:
        """
        # for e in lst:
        # i += 1
        # categ_id = int(self._cats[item['cat_name']])

        # create_categories
        # cat_name = item['cat_name']

        cat_id = self.create_category(item['cat'])
        sname = item['sname']

        # add attributes (only if have his)
        attrs_lines = []
        try:
            for ekey, evalue in item['attributes'].items():
                self.create_attribute(item, attrs_lines, ekey, evalue)
                pass
        except KeyError:
            # pass
            print('item: %s, dont have attributes!' % sname)

        # add variants
        variants_lines = []

        # kind of variant
        cat_to_variants_site_name_value = 'Розмір'
        cat_to_variants_site_name = {
            'Дивани': 'Кут',
            'Ліжка': 'Розмір спального місця',
            'Стільці та табурети': 'Колір',
            'Столи гостьові': 'Колір',
            'Столи журнальні': 'Колір',
            # '': '',
        }

        try:
            cat_to_variants_site_name_value = cat_to_variants_site_name[item['cat']]
        except KeyError:
            pass

        try:
            for variant_name, variant_price in item['variant'].items():
                out('    [%s] = %i' % (variant_name, variant_price))
                self.create_attribute(item, variants_lines, cat_to_variants_site_name_value, variant_name,
                                      variant_price)
                # self.create_attribute(item, variants_lines, 'Розмір', variant_name, variant_price)
        except KeyError:
            # pass
            out('         dont have variants: [%s]' % sname)

        images_array = item['images']

        # main image
        image64 = str(images_array[0]) if len(images_array) > 0 else ''

        # additional images
        product_image_ids = []

        if len(images_array) > 1:
            for image in images_array[1:]:
                add_image = [
                    0,
                    False,
                    {
                        'image': image,
                        'name': sname,
                        'product_tmpl_id': False,
                    }
                ]
                product_image_ids.append(add_image)

        original_variant_list = copy.deepcopy(variants_lines)

        # add variants attributes to attrr_lines
        if len(variants_lines) > 0:
            # l2 = []
            l2 = variants_lines[0]  # initial line
            tmp_list = [x[2]['value_ids'][0] for x in variants_lines]

            for variant in tmp_list:
                l2[2]['value_ids'].append(variant)

            attrs_lines.append(l2)

        product_template_id = self._models_objects.execute_kw(
            self._db, self._uid, self._password,
            'product.template', 'create',
            [{
                'name': sname + ' - 44',
                'list_price': item['price'],
                'company_id': 1,
                'sale_ok': True,
                'purchase_ok': False,
                # 'categ_id': 6,  # All / Можна продавати / Physical  (odoo 10)
                'categ_id': 1,  # All / Salable (odoo 11)
                'image_medium': image64,  # main image
                'attribute_line_ids': attrs_lines,
                'taxes_id': [],
                'supplier_taxes_id': [],

                'product_image_ids': product_image_ids,

                # 'default_code': '1111',
                'public_categ_ids': [
                    [
                        6,
                        False,
                        [cat_id]
                    ]
                ],
                # 'description_sale': 'super_puper_long',

                'website_style_ids': [[
                    6,
                    False,
                    [
                        # 1,
                        # 2
                    ]
                ]],

                'website_description': item['description'],
                'website_published': True,

            }]
        )

        # assign price for variant
        # http://joxi.ru/RmzQMg9T0PYNMr
        extra_price_list = []
        counter = 0
        try:
            for key, elem in item['variant'].items():
                el = original_variant_list[counter][2]

                v = (product_template_id, el['attribute_id'], el['value_ids'][0][1], elem)
                extra_price_list.append(v)
                counter += 1

            self.set_price_for_variants(extra_price_list)
        except KeyError:
            pass

        return product_template_id

    # (product_template_id, attrinute_id, value_id, extra_price)
    def set_price_for_variants(self, extra_price_list):
        # return None

        result = []
        for e in extra_price_list:
            r = self._models_objects.execute_kw(
                self._db, self._uid, self._password,
                'product.attribute.value', 'write',

                [[e[2]],
                 {'price_extra': e[3]}
                 ],

                {
                    'context': {
                        'active_id': e[0],
                        'active_ids':
                            [e[0]],
                        'default_product_tmpl_id': e[0],

                    }

                },

            )
            result.append(r)

        return result

    def set_attributes_for_item(self, id_item, attributes_list):
        id_item = 42
        attributes_list = [
            [{'attribute_id': 7,
              'values_ids': [9]}],
            # [],
        ]

        product_id = self._models_objects.execute_kw(
            self._db, self._uid, self._password,
            'product.template',
            'write',
            [65,
             {'attribute_line_ids': [
                 [
                     0,
                     False,
                     {'attribute_id': 1,
                      'value_ids': [
                          6,
                          False,
                          [1]
                      ]
                      }
                 ],
             ]}],
            # [],
            #    'attribute_line_ids': attributes_list,
        )

        pass

    def test(self):
        # just see in Firefox + F12 debug, template + method + params

        product = self._models_objects.execute_kw(
            self._db, self._uid, self._password,
            'product.template', 'read',
            [
                # 3,
                10,
                # product_id,
            ]
        )[0]

        pass


IMAGES_DOMAINE = 'https://images.ua.prom.st/'
# region 'conditions list const'
CON_IGNORE_SNAME_STARTS = ['Тканини для асортименту',
                           ]

CON_IGNORE_ATTRS = ['Гарантийный срок',
                    'Ручки для переноса',
                    'Тип подъемного механизма',
                    'Количество зон жесткости матраса',
                    'Количество спальных мест',
                    'Вид кровати',
                    'Эффект "Зима - Лето"',
                    'Размер матраса, см',
                    ]


# endregion


class Impxls(object):
    _flush = False
    _add_images = False
    _rebuild_index = False
    _csv1 = None
    _csv2 = None
    _images_folder = '%s\images\\' % os.getcwd()
    _url_lib_pool = urllib3.PoolManager()
    _url_lib_pool.headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:57.0) Gecko/20100101 Firefox/57.0',

        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'ru,uk;q=0.8,en-US;q=0.5,en;q=0.3',
        'Accept-Encoding': 'gzip, deflate, br',

        'Referer': 'https://svitkomforty.com.ua/site_search?search_term=%D0%9A%D0%BE%BD%D1%81%22+',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
    }

    def __init__(self, flush=False, add_images=False, rebuild_index=False):
        self._flush = flush
        self._add_images = add_images
        self._rebuild_index = rebuild_index
        self._csv1 = open('attr1.csv', 'w')
        self._csv2 = open('attr2.csv', 'w')

    def handle(self, fn):
        # def translate(id_kind, text):
        #     result = text
        #     return result
        wb2 = load_workbook(fn, read_only=True)
        # print('Spread sheats names: %s' % wb2.get_sheet_names())
        wb = wb2.worksheets[0]

        index = 0
        cats = dict()
        attr_dict = dict()
        brands = {}

        extra_attr_dict = dict()  # name, [type, value]

        total_count = wb.max_row

        item = {}

        error_attrbitutes_values = {}
        attrs_list_names = {}

        result = []

        for row in wb.rows:
            extra_attr_dict.clear()
            # region 'work'
            index += 1

            if index == 1:
                continue

            # if index < 2738:
            #     continue

            # if index > 50:
            #     break

            item['index'] = index - 1
            item['group_id'] = str(row[28].value).strip()
            item['unique_id'] = str(row[20].value).strip()

            images = str(row[11].value).split(',')
            item['images'] = []
            for e in images:
                # e = e.strip()
                item['images'].append(e.replace(IMAGES_DOMAINE, '').replace('_w640_h640_', '_w5000_h5000_').strip())

                # fdata = get_image_base64_from_url(self._url_lib_pool, e)
                # item['images'].append(fdata)
            # item['images'] = str(row[11].value).replace(IMAGES_DOMAINE, '').split(',')

            # name
            item['sname'] = str(row[1].value).strip() \
                .replace('"', '') \
                .replace('  ', ' ') \
                .replace('*', 'x')

            # ignore some "with start" sname products
            for e in CON_IGNORE_SNAME_STARTS:
                if item['sname'].startswith(e):
                    item['sname'] = ''
            if item['sname'] == '':  # hack
                continue

            out('[%i/%i] %s' % (index, total_count, item['sname']))

            v = row[5].value
            item['price'] = int(v) if not (v is None) else 0

            # v = row[3].value

            item['original_cat'] = str(row[15].value)

            cat = item['original_cat'] \
                .replace('Матраци Sleep&Fly', 'Матраци') \
                .replace('Матраци Evolution', 'Матраци') \
                .replace('Матраци Sleep&fly Organic', 'Матраци') \
                .replace('Матраци Take&Go Bamboо', 'Матраци') \
                .replace('Take&Go', 'Матраци') \
                .replace('Матраци Sleep&fly uno', 'Матраци') \
                .replace('Матраци Doctor Health', 'Матраци') \
                .replace('Дитячі матраци Herbalis KIDS', 'Матраци') \
                .replace('Матраци на дивани', 'Футони і топери') \
                .replace('Наматрацники', 'Наматрацники і підматрацники') \
                .replace("Дерев'яні ліжка", 'Ліжка') \
                .replace("Дитячі ліжка", 'Ліжка') \
                .replace('Столи', 'Столи гостьові') \
                .replace('Столи гостьові-трансформери', 'Столи журнальні') \
                .replace('Стільці', 'Стільці та табурети') \
                .replace('Дитячі дивани', 'Дивани') \
                .replace('Кутові дивани', 'Дивани') \
                .replace('Прямі дивани', 'Дивани') \
                .strip()

            item['cat'] = cat
            cats[cat] = None

            # brand
            tmp = str(row[24].value).strip()
            tmp = tmp \
                .replace('Скиф', 'Скіф') \
                .replace('Тиса мебель', 'Тиса меблі') \
                .replace('Елисеевская мебель', 'Єлисеївські меблі') \
                .replace('Микс мебель', 'Мікс меблі') \
                .replace('Мелитополь мебель', 'Мелітополь меблі') \
                .replace('None', 'Олімп')

            # .replace('Еврокнижка', 'Єврокнижка') \
            # .replace('деревянные ламели', 'букові ламелі')

            extra_attr_dict['Бренд'] = [item['sname'], tmp, 'str', '']

            # country
            tmp = str(row[26].value).strip()
            tmp = tmp.replace('Украина', 'Україна')
            if tmp == '':
                tmp = 'Україна'
            extra_attr_dict['Країна виробник'] = [item['sname'], tmp, 'str', '']

            some_list = list(row[30:])

            # add garanty 18 month attribute
            if cat in [
                'Матраци',
                'Дивани',
                'Подушки',
                'Наматрацники і підматрацники',
                'Футони і топери',
            ]:
                # extra_attr_dict['Гарантыя'] = '18'  # просто добавляем в словарь без типов данных
                extra_attr_dict['Гарантія'] = [item['sname'], '18', 'int', 'міc']

            # use step by 1

            # region 'convert work'
            attrs_list0 = [[item['sname'],
                            str(val.value).strip(),
                            str(some_list[idx + 1].value).strip(),
                            str(some_list[idx + 2].value).strip()
                            .replace('.0', '')
                            .replace('*', 'x')
                            ]
                           for idx, val in enumerate(some_list)
                           if idx % 3 == 0
                           and val.value is not None]

            # skip ignored attributes
            attrs_list1 = [e for e in attrs_list0
                           if e[1] not in CON_IGNORE_ATTRS
                           and not (e[1] == 'Цвет' and e[3] == 'Разные цвета')
                           and not (e[1] == 'Розмір' and e[3] == '7000')
                           and not (e[1] == 'Состояние' and e[3] == 'Новое')
                           and not (e[1] == 'Тип' and e[3] == 'Для сна')
                           and not (e[1] == 'Цвет' and e[3] == 'Белый')
                           and not (e[1] == 'Цвет' and e[3] == 'Разные цвета')
                           and not (e[1] == 'Цвет обивки' and e[3] == 'Разные цвета')
                           and not (e[1] == 'Тип крепления к матрасу' and e[3] == 'четыре резинки по углам')
                           ]

            for e in attrs_list1:
                attrs_list_names[e[1]] = ''

            # mm to sm
            for idx, e in enumerate(attrs_list1):
                if cat in ['Столи гостьові',
                           'Стільці та табурети', ] \
                        and e[1] in ['Глубина столика',
                                     'Длина столика',
                                     'Максимальная длина столешницы раскладного столика',
                                     'Минимальная длина столешницы раскладного столика',
                                     'Длина стола',
                                     'Высота',
                                     'Высота стола',
                                     'Длина стола в раздвинутом (разложенном) состоянии',
                                     'Длина стола в сдвинутом (сложенном) состоянии',
                                     'Ширина',
                                     'Ширина стола',
                                     'Глубина',
                                     ]:
                    e[3] = str(int(e[3]) / 10).replace('.0', '')
                    e[2] = 'см'

                if e[2] in ['см', 'кг', 'шт.']:
                    e[2] = 'int'

                if e[3] in ['да', 'нет', 'True', 'False']:
                    e[2] = 'bool'

                    e[3] = e[3].replace('да', 'так').replace('нет', 'ні').replace('True', 'так').replace('False', 'ні')

                # else string
                # if e[2] == '':
                #    e[2] = 'str'

                # attrs_list2.append(result)

            attrs_list = attrs_list1

            # add extra data
            for e in extra_attr_dict:
                attrs_list.append([item['sname'], e, extra_attr_dict[e][0], extra_attr_dict[e][1]])
                # attrs_list.append([e, extra_attr_dict[e][1]])

            # item['attributes'] = attrs_list
            item['attributes'] = {}
            for e in attrs_list:
                e[3] = e[3].strip().replace('  ', ' ')
                # skip empty attribute values
                if not e[3] == '':  # !!!!!!!!!!!!!!!
                    item['attributes'][e[1]] = e[3]
                    # test anomality attributes values !!!
                    if e[3] in [
                        '', 'г', '7', '309',
                        '890', '760',
                        # '4.7', '4.5', '7.6',  # PROM noobs
                        'False', 'True',
                        'Tik-Tak', 'взаимозаменяемый', 'левый',
                        'одноярусная кровать'
                        # '101', '102', '103', '104', '105', '106', '107', '108',
                    ]:
                        error_attrbitutes_values[e[0] + ', ' + e[1] + '=' + e[3]] = ''
                        out('      ******     Error attribute value: %s      *********' % e[3])

                    # translate
                    # attr names
                    # e[1] = translate(0, e[1])

                    # attr values
                    # e[3] = translate(1, e[3])
                    # pass

                    # error_attrbitutes_values[e[0] + ', ' + e[1] + '=' + e[3]] = ''
                    # out('empty attribute value')
            # detect types of data

            # endregion

            # OUTPUT
            for e in attrs_list:
                s = '%s\t%s\t%s\t%s\n' % (e[0], e[1], e[2], e[3])
                self._csv1.writelines(s)

                if e[3].strip() == '':
                    a = 1
                # just for debug
                attr_dict[e[3]] = ''
                # self._csv += e[0]+'\t'+e[1]+'\t'+e[2]+'\n'

            # print('[%i/%i] [%s] %s' % (index - 1, wb.max_row, cat, row[1].value,))

            item['description'] = str(row[3].value) if not (row[3].value is None) else ''
            result.append(copy.copy(item))

        self._csv1.close()
        self._csv2.close()
        wb2.close()

        out('\n\n==Cats==')
        for e in cats:
            out(e)

        # attrs_list_names
        out('\n\n==Attr_names_dict==')
        for e in sorted(attrs_list_names):
            out(e)

        out('\n\n==Attr_values_dict==')
        for e in sorted(attr_dict):
            out(e)

        out('\n\n==Error: error_attrbitutes_values==')
        for e in sorted(error_attrbitutes_values):
            out(e)

        return result

    @staticmethod
    def assign_main_variant(el):
        el_extra = {}
        s0 = '190/200'

        variants = el['variant']

        for e in variants:

            # 160*190/200 -> 160*190 + 160*200
            if s0 in e:
                price = variants[e]
                prefix = e[len(s0):]
                s1 = e.replace(s0, '190')
                s2 = e.replace(s0, '200')

                el_extra[s1] = price
                el_extra[s2] = price
                # pass
            else:
                el_extra[e] = variants[e]

        key_error = False
        l1 = ['Матраци', 'Ліжка']
        l2 = ['Футони і топери']
        # assign most popular attribute to main record
        if len(el['variant']) > 0:
            # has variants ?
            if el['cat'] in l1:
                try:
                    el['price'] = el_extra['160x200']
                except KeyError:
                    key_error = True

            if el['cat'] in l2:
                try:
                    el['price'] = el_extra['120x190']
                except KeyError:
                    key_error = True

        l3 = l1 + l2
        if key_error or (el['cat'] not in l3):
            # get minimum price from variants
            key_min = min(el_extra.keys(), key=(lambda k: el_extra[k]))

            min_price = el_extra[key_min]
            el['price'] = min_price

            out('*** dont have 160x200 *** [%s] [base_price=%i] %s' % (el['sname'], min_price, el_extra))

        # fix price, from total price to base + extra ptice
        for e3 in el_extra:
            price_base = el['price']
            price_variant = el_extra[e3]
            el_extra[e3] = price_variant - price_base

        el['variant'] = el_extra

        # el['sname'] += 'tetete'
        pass

    def stage2(self, data):
        # dict of group_id with miniumal price
        tmp_list_1 = {}

        for e in data:
            # hack, fix wrong value represent
            if e['group_id'] == 'None':
                e['group_id'] = None

            group_id = e['group_id']

            # counter = 1

            # this is group ? (can be once position!), dict(group_id, sname_with_minimal_length)
            if group_id:
                # v = ''
                try:
                    v = tmp_list_1[group_id]  # do not Delete!!! (need for try test!)
                    len_sname = len(e['sname'])
                    # tmp_list_1[group_id] = tmp_list_1[group_id] + counter  # new group_id, append!

                    if len_sname < len(e['sname']):
                        tmp_list_1[group_id] = len_sname  # new group_id, append!
                except KeyError:
                    # tmp_list_1[group_id] = counter  # new group_id, append!
                    tmp_list_1[group_id] = len(e['sname'])

        tmp_list_3 = {}

        # assign main position
        out('\n==Base positions (for variants)==')
        for e in data:
            group_id = e['group_id']
            if group_id:
                v = tmp_list_1[group_id]
                if len(e['sname']) == v:
                    tmp_list_3[group_id] = e
                    tmp_list_3[group_id]['variant'] = dict()
                    out('%s: %i' % (e['sname'], e['price']))

        # add variants
        for e in data:
            group_id = e['group_id']
            sname_len = len(e['sname'])
            if group_id and (sname_len > len(tmp_list_3[group_id]['sname'])):
                # add variant
                e2 = tmp_list_3[group_id]

                sname_len2 = len(e2['sname'])
                s = e['sname'][sname_len2:].strip()  # [sname_len::]

                # dict(variant, price)
                e2['variant'][s] = e['price']

                _ = 1
                # prom.ua create separate image of variants for same item 0_o.... world - stop!, i live this planete...
                """
                for image in e['images']:
                    if image not in e2['images']:
                        e2['images'].append(image)
                """

        # assign variant to main variant
        # tmp_list_4 = {}
        for e in tmp_list_3:
            item = tmp_list_3[e]
            self.assign_main_variant(item)
            if item['attributes']['Бренд'] == '':
                out('***brand empty!: %s***' % e)
                raise Exception('brand empty! = 0', 'brand')
            if item['price'] == '0':
                # out('***price 0!: %s***' % e)
                raise Exception('[%s] price = 0' % e, 'price')

        return tmp_list_3

    def stage3(self, data_variants, data_full):
        # cls()

        tmp_list = []
        for e in data_variants:
            tmp_list.append(data_variants[e])

        for e in data_full:
            if not e['group_id']:
                tmp_list.append(e)

        # cache images
        counter = 0
        for e in tmp_list:
            counter += 1

            # parse page for images (real and big!!!, not resized!)
            # url = 'https://svitkomforty.com.ua/p%s-prodam-garaj.html' % e['unique_id']
            # url = 'https://svitkomforty.com.ua/p595800165-prodam-garaj.html'
            # url = 'https://svitkomforty.com.ua/p595800165-komod-provans.html'
            # url = 'https://pgl.yoyo.org/http/browser-headers.php'
            # html_data = ''
            #
            # with self._url_lib_pool.request('GET', url, preload_content=False) as resp:
            #     html_data = resp.read()
            #
            #     with open('file.html', 'wb') as f:
            #         f.write(html_data)

            if e['sname'] == 'Ліжко Далі Люкс':
                _ = 1

            new_images = []

            for svalue in e['images']:
                if len(svalue) < 2:
                    continue

                fs = '%s%s' % (self._images_folder, svalue)
                if not os.path.isfile(fs):
                    url = '%s%s' % (IMAGES_DOMAINE, svalue)

                    with self._url_lib_pool.request('GET', url, preload_content=False) as resp, \
                            open(fs, 'wb') as out_file:
                        shutil.copyfileobj(resp, out_file)

                    resp.release_conn()  # not 100% sure this is required though
                    out('[%i/%i] download: %s' % (counter, len(tmp_list), url))

                if not Utils.check_image_correct(fs):
                    out('            ***Error: image open fail: %s, %s ***' % (e['sname'], fs))
                    # e['images'][skey] = ''
                else:
                    new_images.append(svalue)
                # TODO: crop white spaces

            e['images'] = new_images
        _ = 1

        return tmp_list

    def stage10(self, import_object, data):
        counter = 0
        for e in data:
            counter += 1
            # process images
            images = e['images']
            images64 = []
            for image in images:
                if len(image) > 0:
                    images64.append(Utils.get_base('%s/%s' % (self._images_folder, image)))

            e['images'] = images64
            out('[%i/%i]: [id: %i] [%s] ' % (counter, len(data), e['index'], e['sname']))
            import_object.create_item(e)
            # pass


# c = Impxls()
# c.handle('export-products.xlsx')

# test
im = ImportToOdd(env('host'), env('db'), env('user'), env('pwd'))
# im.connect()

# im.test()
# im.unlink_attributes()
# im.create_attribute('форма', 'квадрат')

# im.set_attributes_for_item(None, None)
# cats = im.create_categories(list_1)

# im.unlink_item()

c = Impxls(True, True)

# stage 1, prepare XLS to import
# data = None

# TODO: NEED UNKOMENT!!!!!
"""
with open('stage1.pickle', 'wb') as handle:
    pickle.dump(data_xls, handle, protocol=pickle.HIGHEST_PROTOCOL)
with open('stage1.pickle', 'rb') as handle:  # stage 2, combine items to variants group,
    data_xls = pickle.load(handle)

with open('stage2.pickle', 'wb') as handle:
    pickle.dump(data_variants2, handle, protocol=pickle.HIGHEST_PROTOCOL)
with open('stage2.pickle', 'rb') as handle:  # stage 3, cat + brand = attributes + variants
    data = pickle.load(handle)

with open('stage3.pickle', 'wb') as handle:
    pickle.dump(data, handle, protocol=pickle.HIGHEST_PROTOCOL)    
"""

# exit(0)

data_xls = c.handle('export-products.xlsx')
data_variants2 = c.stage2(data_xls)
ready_data = c.stage3(data_variants2, data_xls)

c.stage10(im, ready_data)
